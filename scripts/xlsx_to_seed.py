#!/usr/bin/env python3
"""Convert DAS Construction Bidding Workbook.xlsx into seed JSON files.

Run: python scripts/xlsx_to_seed.py
Reads:  ./DAS Construction Bidding Workbook.xlsx
Writes: ./prisma/seed-data/*.json
"""
import json
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "DAS Construction Bidding Workbook.xlsx"
OUT_DIR = ROOT / "prisma" / "seed-data"

ROLE_NAMES = [
    "Technician", "Construction Manager", "RF-Engineer",
    "RF-Technician", "Project Coordinator", "Project Manager",
]

# LOE Sheet rows whose Quantity (column A) is itself a formula (a derived
# quantity) rather than a direct user input. Resolved by hand from the
# workbook's actual formula text (verified against the source file) into a
# linear combination: quantity = (sum of coeff * source_row_quantity) / divisor.
#   loe-25 = SUM(A21:A24, A31)         -> labeling coax/category cable
#   loe-28 = A27*4                     -> labeling for splitter
#   loe-32 = A31/2                     -> test category cable per drop
#   loe-33 = SUM(A21:A24)/2            -> sweep test per line
#   loe-46 = SUM(A43:A44)*2 - A43      -> labeling fiber (= A43 + 2*A44)
#   loe-49 = A48+A47                   -> labeling fiber housing
#   loe-67 = SUM(A65,A64,A63,A62,A56,A57) -> labeling DAS equipment
#   loe-71 = SUM(A83,A72,A73,A75,A76,A77,A78,A79,A82,A70) -> labeling grounding
LOE_DERIVED_QUANTITIES = {
    25: {"terms": [(21, 1), (22, 1), (23, 1), (24, 1), (31, 1)], "divisor": 1},
    28: {"terms": [(27, 4)], "divisor": 1},
    32: {"terms": [(31, 1)], "divisor": 2},
    33: {"terms": [(21, 1), (22, 1), (23, 1), (24, 1)], "divisor": 2},
    46: {"terms": [(43, 1), (44, 2)], "divisor": 1},
    49: {"terms": [(48, 1), (47, 1)], "divisor": 1},
    67: {"terms": [(65, 1), (64, 1), (63, 1), (62, 1), (56, 1), (57, 1)], "divisor": 1},
    71: {"terms": [(83, 1), (72, 1), (73, 1), (75, 1), (76, 1), (77, 1),
                    (78, 1), (79, 1), (82, 1), (70, 1)], "divisor": 1},
}

SUM_RANGE_RE = re.compile(r"SUM\(G(\d+):G(\d+)\)")


def load_wb():
    return openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)


def load_wb_values():
    return openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)


def parse_material_items(wb):
    ws = wb["Bill of Materials"]
    items = []
    for row in range(3, 87):
        item_type = ws[f"A{row}"].value
        category = ws[f"G{row}"].value
        unit_cost = ws[f"H{row}"].value
        if item_type is None or category is None or unit_cost is None:
            continue
        items.append({
            "key": f"bom-{row}",
            "type": item_type,
            "manufacturer": ws[f"B{row}"].value,
            "model": str(ws[f"C{row}"].value) if ws[f"C{row}"].value is not None else None,
            "description": ws[f"D{row}"].value,
            "vendor": ws[f"F{row}"].value,
            "category": category,
            "unitCost": float(unit_cost),
        })
    return items


def parse_labor_sheet(wb, sheet_name, key_prefix, derived_quantities):
    ws = wb[sheet_name]
    tasks = []
    current_category = None
    current_subtotal_range = None
    for row in range(3, ws.max_row + 1):
        marker = ws[f"A{row}"].value
        if marker == "~":
            current_category = ws[f"B{row}"].value
            # Skip category markers without a name (these are summary rows)
            if current_category is None:
                continue
            g_formula = ws[f"G{row}"].value
            match = SUM_RANGE_RE.search(g_formula) if isinstance(g_formula, str) else None
            if not match:
                raise ValueError(
                    f"{sheet_name}!G{row}: expected a SUM(G#:G#) subtotal formula, got {g_formula!r}"
                )
            current_subtotal_range = (int(match.group(1)), int(match.group(2)))
            continue
        name = ws[f"B{row}"].value
        minutes = ws[f"C{row}"].value
        unit = ws[f"D{row}"].value
        role = ws[f"E{row}"].value
        if name is None or role is None:
            continue
        # Skip header rows that have "Labor Type" as the role label
        if role == "Labor Type":
            continue
        if role not in ROLE_NAMES:
            raise ValueError(f"{sheet_name}!E{row}: unrecognized labor role {role!r}")
        derived = derived_quantities.get(row)
        derived_out = None
        if derived is not None:
            derived_out = {
                "terms": [{"key": f"{key_prefix}-{r}", "coeff": c} for r, c in derived["terms"]],
                "divisor": derived["divisor"],
            }
        included = (
            current_subtotal_range is not None
            and current_subtotal_range[0] <= row <= current_subtotal_range[1]
        )
        tasks.append({
            "key": f"{key_prefix}-{row}",
            "sheet": "LOE" if key_prefix == "loe" else "SOW",
            "category": current_category,
            "name": name,
            "minutesPerUnit": float(minutes) if isinstance(minutes, (int, float)) else 0.0,
            "unit": unit if isinstance(unit, str) and unit else "Each",
            "laborRole": role,
            "includedInSubtotal": included,
            "derivedFrom": derived_out,
        })
    return tasks


def parse_labor_rates(wb_values):
    # NOTE: must be called with a workbook loaded via data_only=True. Column D
    # ("Reg Bill with MU") is a formula for 5 of 6 roles and a hardcoded
    # literal override for RF-Engineer (=100, vs its raw wage B5=75) — reading
    # it with data_only=False would return the formula text, not the number.
    ws = wb_values["Labor Projections"]
    rates = []
    for row in range(3, 9):
        role = ws[f"A{row}"].value
        raw_wage = ws[f"B{row}"].value
        billing_rate = ws[f"D{row}"].value
        if role is None or raw_wage is None or billing_rate is None:
            continue
        rates.append({
            "role": role,
            "hourlyRate": float(billing_rate),
            "rawWageRate": float(raw_wage),
        })
    return rates


def parse_crew_size_table(wb):
    ws = wb["Labor Projections"]
    rows = []
    for row in range(3, 23):
        tech_count = ws[f"I{row}"].value
        cms_needed = ws[f"L{row}"].value
        if tech_count is None or cms_needed is None:
            continue
        rows.append({"technicianCount": int(tech_count), "cmsNeeded": int(cms_needed)})
    return rows


def parse_labor_projection_settings(wb):
    ws = wb["Labor Projections"]
    return {
        "hoursPerManDay": float(ws["B22"].value),
        "hoursPerManWeek": float(ws["B24"].value),
        "stagingMaterialMultiplier": float(ws["B18"].value),
        "cmPercentOfTechHours": float(ws["E11"].value),
        "pmPercentOfTechHours": float(ws["E12"].value),
        "coordinatorPercentOfTechHours": float(ws["E13"].value),
    }


def parse_pass_through_rates(wb_values):
    # NOTE: must be called with the data_only=True workbook. Column A (role
    # name) is a cross-sheet formula (e.g. ='Labor Projections'!A$3) for the
    # perDiem/lodging/airfare role rows; reading it with data_only=False
    # would return the formula text, not the resolved role name.
    ws = wb_values["Pass Throughs"]

    def role_rate_rows(rows, value_col):
        out = []
        for row in rows:
            role = ws[f"A{row}"].value
            value = ws[f"{value_col}{row}"].value
            if role is not None and value is not None:
                out.append({"role": role, "value": float(value)})
        return out

    per_diem = role_rate_rows(range(5, 11), "B")
    lodging = role_rate_rows(range(15, 21), "B")
    airfare = role_rate_rows(range(35, 41), "B")

    rentals = []
    for row in range(45, 58):
        name = ws[f"A{row}"].value
        rate = ws[f"B{row}"].value
        unit = ws[f"C{row}"].value
        if name is None or rate is None:
            continue
        rentals.append({"key": f"rental-{row}", "name": name, "rate": float(rate), "unit": unit or "Each"})

    soft_costs = []
    for row in range(62, 72):
        name = ws[f"A{row}"].value
        fee = ws[f"D{row}"].value
        if name is None:
            continue
        soft_costs.append({"key": f"softcost-{row}", "name": name, "fee": float(fee) if fee is not None else 0.0})

    return {
        "perDiemRateByRole": [{"role": r["role"], "rate": r["value"]} for r in per_diem],
        "lodgingRateByRole": [{"role": r["role"], "rate": r["value"]} for r in lodging],
        "airfareCostByRole": [{"role": r["role"], "cost": r["value"]} for r in airfare],
        "rentals": rentals,
        "softCosts": soft_costs,
    }


def main():
    wb = load_wb()
    wb_values = load_wb_values()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    material_items = parse_material_items(wb)
    assert len(material_items) >= 80, f"expected ~84 material items, got {len(material_items)}"

    loe_tasks = parse_labor_sheet(wb, "LOE Sheet", "loe", LOE_DERIVED_QUANTITIES)
    sow_tasks = parse_labor_sheet(wb, "Additional SOW's", "sow", {})
    assert len(loe_tasks) == 89, f"expected exactly 89 LOE tasks, got {len(loe_tasks)}"
    assert len(sow_tasks) == 19, f"expected exactly 19 Additional SOW tasks, got {len(sow_tasks)}"
    labor_tasks = loe_tasks + sow_tasks

    labor_rates = parse_labor_rates(wb_values)
    assert len(labor_rates) == 6, f"expected 6 labor rates, got {len(labor_rates)}"

    crew_size_table = parse_crew_size_table(wb)
    assert len(crew_size_table) == 20, f"expected 20 crew-size rows, got {len(crew_size_table)}"

    labor_projection_settings = parse_labor_projection_settings(wb)
    pass_through_rates = parse_pass_through_rates(wb_values)

    def write(name, data):
        with open(OUT_DIR / name, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        count = len(data) if isinstance(data, list) else "object"
        print(f"wrote {name}: {count}")

    write("material-items.json", material_items)
    write("labor-tasks.json", labor_tasks)
    write("labor-rates.json", labor_rates)
    write("crew-size-table.json", crew_size_table)
    write("labor-projection-settings.json", labor_projection_settings)
    write("pass-through-rates.json", pass_through_rates)


if __name__ == "__main__":
    main()
